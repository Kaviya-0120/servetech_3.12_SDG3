#!/usr/bin/env python3
"""
Report Generation Module for Hospital System
Generates PDF and Excel reports for patient data
"""

import io
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ReportGenerator:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        
    def generate_patient_pdf_report(self, patient_data):
        """Generate PDF report for a single patient"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#1565c0')
        )
        story.append(Paragraph("MediCare Hospital - Patient Report", title_style))
        story.append(Spacer(1, 12))
        
        # Patient Information Table
        patient_info = [
            ['Patient Information', ''],
            ['Registration ID', patient_data.get('registration_id', 'N/A')],
            ['Name', patient_data.get('patient_name', 'N/A')],
            ['Age', str(patient_data.get('age', 'N/A'))],
            ['Gender', patient_data.get('gender', 'N/A')],
            ['Phone', patient_data.get('phone', 'N/A')],
            ['Department', patient_data.get('department', 'N/A')]
        ]
        
        patient_table = Table(patient_info, colWidths=[2*inch, 3*inch])
        patient_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(patient_table)
        story.append(Spacer(1, 20))
        
        # Medical Information
        medical_info = [
            ['Medical Assessment', ''],
            ['Primary Symptom', patient_data.get('symptom', 'N/A')],
            ['Severity Level', patient_data.get('severity', 'N/A')],
            ['Risk Score', f"{patient_data.get('risk_score', 'N/A')}/100"],
            ['Emergency Case', 'Yes' if patient_data.get('is_emergency') else 'No'],
            ['Status', patient_data.get('status', 'N/A').title()],
            ['Registration Date', patient_data.get('created_at', 'N/A')],
            ['Appointment Time', patient_data.get('appointment_time', 'Not Scheduled')]
        ]
        
        medical_table = Table(medical_info, colWidths=[2*inch, 3*inch])
        medical_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f57c00')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(medical_table)
        story.append(Spacer(1, 20))
        
        # Footer
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer_style = ParagraphStyle(
            'Footer',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.grey
        )
        story.append(Paragraph(footer_text, footer_style))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_daily_report_pdf(self, report_data):
        """Generate daily operations PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            textColor=colors.HexColor('#1565c0'),
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Daily Operations Report", title_style))
        story.append(Paragraph(f"Date: {report_data.get('date', 'N/A')}", self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary Statistics
        summary_data = [
            ['Daily Summary', ''],
            ['Total Patients', str(report_data.get('total_patients', 0))],
            ['High-Risk Cases', str(report_data.get('high_risk_patients', 0))],
            ['Confirmed Appointments', str(report_data.get('confirmed_appointments', 0))],
            ['Report Generated', report_data.get('generated_at', 'N/A')]
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Department Breakdown
        if report_data.get('department_breakdown'):
            dept_data = [['Department', 'Patient Count']]
            for dept, count in report_data['department_breakdown'].items():
                dept_data.append([dept, str(count)])
            
            dept_table = Table(dept_data, colWidths=[2.5*inch, 2*inch])
            dept_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#388e3c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(Paragraph("Department Breakdown", self.styles['Heading2']))
            story.append(dept_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_patients_excel_report(self, patients_data):
        """Generate Excel report with all patients data"""
        buffer = io.BytesIO()
        
        # Create DataFrame
        df = pd.DataFrame(patients_data)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Patients Report"
        
        # Add title
        ws['A1'] = "MediCare Hospital - Patients Report"
        ws['A1'].font = Font(size=16, bold=True, color="1565C0")
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=12, color="666666")
        
        # Add data starting from row 4
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        header_row = 4
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_daily_report_excel(self, report_data):
        """Generate daily report in Excel format"""
        buffer = io.BytesIO()
        wb = Workbook()
        
        # Summary sheet
        ws1 = wb.active
        ws1.title = "Daily Summary"
        
        # Add title and data
        ws1['A1'] = "Daily Operations Report"
        ws1['A1'].font = Font(size=16, bold=True, color="1565C0")
        ws1['A2'] = f"Date: {report_data.get('date', 'N/A')}"
        ws1['A2'].font = Font(size=12)
        
        # Summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Patients", report_data.get('total_patients', 0)],
            ["High-Risk Cases", report_data.get('high_risk_patients', 0)],
            ["Confirmed Appointments", report_data.get('confirmed_appointments', 0)],
            ["Report Generated", report_data.get('generated_at', 'N/A')]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header row
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        
        # Department breakdown sheet
        if report_data.get('department_breakdown'):
            ws2 = wb.create_sheet("Department Breakdown")
            ws2['A1'] = "Department Breakdown"
            ws2['A1'].font = Font(size=14, bold=True, color="1565C0")
            
            dept_data = [["Department", "Patient Count"]]
            for dept, count in report_data['department_breakdown'].items():
                dept_data.append([dept, count])
            
            for row_idx, row_data in enumerate(dept_data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 3:  # Header row
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer